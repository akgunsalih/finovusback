from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from io import BytesIO
from datetime import datetime, date
import json
from typing import List, Optional, Dict
from pydantic import BaseModel
import asyncio
from fastapi import WebSocket, WebSocketDisconnect
import random

app = FastAPI(title="Finovus API")

# Enable CORS for frontend development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- WebSocket & Global State ---
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def lock_connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                continue

manager = ConnectionManager()
latest_data = {"meta": None, "sonuclar": []}
simulation_task = None

# --- Models ---
class SonucRow(BaseModel):
    kontrat: Optional[str]
    aciklama: Optional[str]
    alis: Optional[float]
    spot_satis: Optional[float]
    gun: Optional[int]
    gun_fark: Optional[float]
    spot_gun_fark: Optional[float]
    hesaplama: Optional[float]
    referans_faiz: Optional[float]
    islem_onerisi: Optional[str]

class SpotRow(BaseModel):
    sembol: Optional[str]
    son_fiyat: Optional[float]
    alis: Optional[float]
    satis: Optional[float]
    gun_fark: Optional[float]
    islem_onerisi: Optional[str]

class MetaInfo(BaseModel):
    bugun: Optional[str]
    referans_faiz: Optional[float]
    toplam_satir: int
    islem_yap: int
    islem_yapma: int
    spot_toplam_satir: int
    spot_islem_yap: int
    spot_islem_yapma: int

class CalculationResult(BaseModel):
    meta: MetaInfo
    sonuclar: List[SonucRow]
    spot_sonuclar: List[SpotRow]

# --- Helper Functions (Migrated from finovus_hesapla.py) ---
def to_date(val) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            # ISO format (2024-04-15 or 2024-04-15T12:00:00)
            return datetime.fromisoformat(val[:10]).date()
        except:
            pass
    return None

def sheet_to_list(ws) -> List[Dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, row)))
    return result

# --- Core Calculation Logic ---
def calculate_from_sheets(sheets_data: Dict[str, List[Dict]]) -> dict:
    # REFERANS FAİZ
    rf_rows = sheets_data.get("REFERANS FAİZ", [])
    referans_faiz = None
    for r in rf_rows:
        if str(r.get("KOD", "")).strip() == "TLREF":
            try:
                referans_faiz = float(r["FAİZ"])
            except:
                pass
            break
            
    # SÖZLEŞME TARİH
    st_rows = sheets_data.get("SÖZLEŞME TARİH", [])
    sozlesme = {}
    for r in st_rows:
        key = str(r.get("TARİH", "")).strip()
        val = to_date(r.get("VADE SONU"))
        if key and val:
            sozlesme[key] = val
    
    bugun_tarihi = date.today()
    
    # MATRİKS VERİ SPOT
    spot_rows = sheets_data.get("MATRİKS VERİ SPOT", [])
    spot = {}
    spot_sonuclar = []
    
    for r in spot_rows:
        sembol = str(r.get("SEMBOL", "")).strip().upper()
        if not sembol: continue
        
        try:
            son_fiyat = float(r.get("SON FİYAT") or 0)
        except:
            son_fiyat = None
            
        try:
            alis = float(r.get("ALIŞ") or 0)
        except:
            alis = None
            
        try:
            satis = float(r["SATIŞ"] or 0)
            gf = r.get("GÜN FARK %")
            gun_fark = float(gf) if gf is not None else 0.0
            spot[sembol] = {"satis": satis, "gun_fark": gun_fark}
        except:
            satis = None
            gun_fark = 0.0
            
        islem_onerisi = "İŞLEM YAP" if gun_fark > 0 else "İŞLEM YAPMA"
        
        spot_sonuclar.append({
            "sembol": sembol,
            "son_fiyat": son_fiyat,
            "alis": alis,
            "satis": satis,
            "gun_fark": gun_fark,
            "islem_onerisi": islem_onerisi
        })
            
    # MATRİKS VERİ VADELİ
    vadeli_rows = sheets_data.get("MATRİKS VERİ VADELİ", [])
    sonuclar = []
    
    for row in vadeli_rows:
        kontrat = str(row.get("KONTRAT", "") or "").strip()
        aciklama = str(row.get("AÇIKLAMA", "") or "").strip()
        
        try:
            alis = float(row.get("ALIŞ") or 0)
        except:
            alis = None
            
        kelimeler = aciklama.split()
        ilk_kelime = kelimeler[0].upper() if len(kelimeler) >= 1 else ""
        ikinci_kel = kelimeler[1].strip() if len(kelimeler) >= 2 else ""
        
        spot_data = spot.get(ilk_kelime, {"satis": None, "gun_fark": None})
        spot_satis = spot_data["satis"]
        spot_gun_fark = spot_data["gun_fark"]
        
        gun = None
        if ikinci_kel and bugun_tarihi:
            vade_sonu = sozlesme.get(ikinci_kel)
            if vade_sonu:
                delta = (vade_sonu - bugun_tarihi).days
                gun = delta if delta > 0 else None
                
        hesaplama = None
        if alis and spot_satis and spot_satis != 0 and gun and gun != 0:
            hesaplama = ((spot_satis / alis) - 1) / gun * 365
            
        islem_onerisi = None
        if hesaplama is not None and referans_faiz is not None:
            hesaplama_pct = hesaplama * 100
            islem_onerisi = "İŞLEM YAP" if hesaplama_pct > referans_faiz else "İŞLEM YAPMA"
            
        gun_fark_pct = None
        try:
            val = row.get("GÜN FARK %")
            if val is not None:
                gun_fark_pct = float(val)
        except:
            gun_fark_pct = 0.0

        sonuclar.append({
            "kontrat":        kontrat,
            "aciklama":       aciklama,
            "alis":           alis,
            "spot_satis":     spot_satis,
            "gun":            gun,
            "gun_fark":       gun_fark_pct,
            "spot_gun_fark":  spot_gun_fark,
            "hesaplama":      round(hesaplama * 100, 4) if hesaplama is not None else None,
            "referans_faiz":  referans_faiz,
            "islem_onerisi":  islem_onerisi,
        })
        
    islem_yap = sum(1 for r in sonuclar if r["islem_onerisi"] == "İŞLEM YAP")
    islem_yapma = sum(1 for r in sonuclar if r["islem_onerisi"] == "İŞLEM YAPMA")
    
    spot_islem_yap = sum(1 for r in spot_sonuclar if r["islem_onerisi"] == "İŞLEM YAP")
    spot_islem_yapma = sum(1 for r in spot_sonuclar if r["islem_onerisi"] == "İŞLEM YAPMA")
    
    return {
        "meta": {
            "bugun": str(bugun_tarihi) if bugun_tarihi else None,
            "referans_faiz": referans_faiz,
            "toplam_satir": len(sonuclar),
            "islem_yap": islem_yap,
            "islem_yapma": islem_yapma,
            "spot_toplam_satir": len(spot_sonuclar),
            "spot_islem_yap": spot_islem_yap,
            "spot_islem_yapma": spot_islem_yapma
        },
        "sonuclar": sonuclar,
        "spot_sonuclar": spot_sonuclar
    }

def perform_calculation(file_content: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True, keep_links=False)
    
    sheets_data = {}
    for sheet_name in ["REFERANS FAİZ", "SÖZLEŞME TARİH", "MATRİKS VERİ SPOT", "MATRİKS VERİ VADELİ"]:
        if sheet_name in wb.sheetnames:
            sheets_data[sheet_name] = sheet_to_list(wb[sheet_name])
        else:
            raise ValueError(f"{sheet_name} sayfası bulunamadı.")
            
    return calculate_from_sheets(sheets_data)

async def run_simulation():
    global latest_data
    while True:
        await asyncio.sleep(3) # Her 3 saniyede bir güncelle
        if not latest_data["sonuclar"]:
            continue
            
        # Simülasyonu tamamen kapattık. Sadece WebSocket bağlantısını canlı tutar.
        await manager.broadcast(latest_data)

def generate_excel(result: dict) -> BytesIO:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "SONUÇLAR"

    headers = ["KONTRAT", "AÇIKLAMA", "ALIŞ", "SPOT SATIŞ", "SPOT FARK %", "VADELİ FARK %",
               "HESAPLAMA %", "REFERANS FAİZ %", "İŞLEM ÖNERİSİ"]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    yesil = PatternFill("solid", fgColor="C6EFCE")
    kirmizi = PatternFill("solid", fgColor="FFC7CE")
    gri = PatternFill("solid", fgColor="F2F2F2")
    beyaz = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(result["sonuclar"], start=2):
        ws.append([
            r["kontrat"], r["aciklama"], r["alis"], r["spot_satis"],
            r.get("spot_gun_fark"), r.get("gun_fark"), r.get("hesaplama"), r.get("referans_faiz"), r.get("islem_onerisi"),
        ])
        fill = gri if i % 2 == 0 else beyaz
        for cell in ws[i]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            cell.border = border

        onerisi_cell = ws.cell(row=i, column=9)
        if r.get("islem_onerisi") == "İŞLEM YAP":
            onerisi_cell.fill = yesil
            onerisi_cell.font = Font(bold=True, color="276221")
        elif r.get("islem_onerisi") == "İŞLEM YAPMA":
            onerisi_cell.fill = kirmizi
            onerisi_cell.font = Font(bold=True, color="9C0006")

        for col in [3, 4, 7, 8]:
            c = ws.cell(row=i, column=col)
            if c.value is not None:
                c.number_format = '0.0000' if col >= 7 else '0.00'

    for hdr_cell in ws[1]:
        hdr_cell.border = border

    col_widths = [18, 38, 10, 12, 12, 12, 14, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # --- SPOT SONUÇLAR ---
    if result.get("spot_sonuclar"):
        ws_spot = wb_out.create_sheet("SPOT SONUÇLAR")
        spot_headers = ["SEMBOL", "SON FİYAT", "ALIŞ", "SATIŞ", "GÜN FARK %", "İŞLEM ÖNERİSİ"]
        ws_spot.append(spot_headers)

        for cell in ws_spot[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for i, r in enumerate(result["spot_sonuclar"], start=2):
            ws_spot.append([
                r["sembol"], r["son_fiyat"], r["alis"], r["satis"],
                r["gun_fark"], r["islem_onerisi"],
            ])
            fill = gri if i % 2 == 0 else beyaz
            for cell in ws_spot[i]:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                cell.border = border

            onerisi_cell = ws_spot.cell(row=i, column=6)
            if r.get("islem_onerisi") == "İŞLEM YAP":
                onerisi_cell.fill = yesil
                onerisi_cell.font = Font(bold=True, color="276221")
            elif r.get("islem_onerisi") == "İŞLEM YAPMA":
                onerisi_cell.fill = kirmizi
                onerisi_cell.font = Font(bold=True, color="9C0006")
                
            for col in [2, 3, 4]:
                c = ws_spot.cell(row=i, column=col)
                if c.value is not None:
                    c.number_format = '0.00'

        col_widths_spot = [12, 12, 12, 12, 12, 18]
        for i, w in enumerate(col_widths_spot, 1):
            ws_spot.column_dimensions[get_column_letter(i)].width = w
        ws_spot.row_dimensions[1].height = 25
        ws_spot.freeze_panes = "A2"

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output

# --- Endpoints ---
@app.get("/")
async def root():
    return {"status": "ok", "message": "Finovus API is running"}

@app.post("/calculate", response_model=CalculationResult)
async def calculate(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Lütfen geçerli bir Excel dosyası yükleyin.")
    
    try:
        content = await file.read()
        global latest_data, simulation_task
        latest_data = perform_calculation(content)
        
        # VERİ GELDİĞİ AN SİTEYE BAS (Bekleme yapma)
        await manager.broadcast(latest_data)
        
        # Simülasyonu başlat (eğer başlamadıysa)
        if simulation_task is None:
            simulation_task = asyncio.create_task(run_simulation())
            
        return latest_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class SyncData(BaseModel):
    sheets: Dict[str, List[Dict]]

@app.post("/calculate-json", response_model=CalculationResult)
async def calculate_json(data: SyncData):
    try:
        global latest_data, simulation_task
        latest_data = calculate_from_sheets(data.sheets)
        
        # VERİ GELDİĞİ AN SİTEYE BAS
        await manager.broadcast(latest_data)
        
        if simulation_task is None:
            simulation_task = asyncio.create_task(run_simulation())
            
        return latest_data
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export")
async def export(file: UploadFile = File(...)):
    try:
        content = await file.read()
        result = perform_calculation(content)
        excel_file = generate_excel(result)
        
        filename = f"FINOVUS_SONUC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.lock_connect(websocket)
    try:
        # Bağlanınca varsa son veriyi hemen gönder
        if latest_data["sonuclar"]:
            await websocket.send_json(latest_data)
        while True:
            await websocket.receive_text() # Bağlantıyı canlı tutmak için bekler
    except WebSocketDisconnect:
        manager.disconnect(websocket)

class ExportData(BaseModel):
    meta: MetaInfo
    sonuclar: List[SonucRow]
    spot_sonuclar: Optional[List[SpotRow]] = None

@app.post("/export-json")
async def export_json(data: ExportData):
    try:
        excel_file = generate_excel(data.dict())
        filename = f"FINOVUS_CANLI_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
