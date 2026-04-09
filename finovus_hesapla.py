"""
FINOVUS ALGO-VAR - Hesaplama Motoru
====================================
Kullanım:
  python finovus_hesapla.py --input FINOVUS.xlsx
  python finovus_hesapla.py --input FINOVUS.xlsx --output SONUC.xlsx

Çıktı: JSON (stdout)
Hatalar: stderr

Formül:
  HESAPLAMA = ((SATIŞ_spot / ALIŞ_vadeli) - 1) / GÜN * 365

  GÜN = VADE_SONU[ay_adi] - VADE_SONU["Bugün"]
        (SÖZLEŞME TARİH sheet'indeki "Bugün" satırı baz alınır)

İŞLEM ÖNERİSİ:
  HESAPLAMA > FAİZ  → "İŞLEM YAP"
  HESAPLAMA < FAİZ  → "İŞLEM YAPMA"
"""

import sys
import json
import argparse
import openpyxl
from datetime import datetime, date

# ─────────────────────────────────────────────────────────
def load_wb(path: str):
    return openpyxl.load_workbook(path, data_only=True, keep_links=False)

def sheet_to_list(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, row)))
    return result

def to_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

# ─────────────────────────────────────────────────────────
def hesapla(input_path: str) -> dict:
    wb = load_wb(input_path)

    # ── REFERANS FAİZ ──────────────────────────────────────
    rf_rows = sheet_to_list(wb["REFERANS FAİZ"])
    referans_faiz = None
    for r in rf_rows:
        if str(r.get("KOD", "")).strip() == "TLREF":
            try:
                referans_faiz = float(r["FAİZ"])
            except Exception:
                pass
            break

    # ── SÖZLEŞME TARİH ────────────────────────────────────
    st_rows = sheet_to_list(wb["SÖZLEŞME TARİH"])
    sozlesme: dict[str, date] = {}
    for r in st_rows:
        key = str(r.get("TARİH", "")).strip()
        val = to_date(r.get("VADE SONU"))
        if key and val:
            sozlesme[key] = val

    bugun_tarihi: date | None = date.today()  # Gerçek bugünün tarihi (Excel'deki sabit değer yerine)

    # ── MATRİKS VERİ SPOT ─────────────────────────────────
    spot_rows = sheet_to_list(wb["MATRİKS VERİ SPOT"])
    # Sembol → SATIŞ fiyatı
    spot: dict[str, float] = {}
    for r in spot_rows:
        sembol = str(r.get("SEMBOL", "")).strip().upper()
        try:
            satis = float(r["SATIŞ"])
        except Exception:
            continue
        if sembol:
            spot[sembol] = satis

    # ── MATRİKS VERİ VADELİ ───────────────────────────────
    vadeli_rows = sheet_to_list(wb["MATRİKS VERİ VADELİ"])

    sonuclar = []

    for row in vadeli_rows:
        kontrat  = str(row.get("KONTRAT", "") or "").strip()
        aciklama = str(row.get("AÇIKLAMA", "") or "").strip()

        try:
            alis = float(row.get("ALIŞ") or 0)
        except Exception:
            alis = None

        kelimeler = aciklama.split()
        ilk_kelime = kelimeler[0].upper()  if len(kelimeler) >= 1 else ""
        ikinci_kel = kelimeler[1].strip()  if len(kelimeler) >= 2 else ""

        # SPOT SATIŞ
        spot_satis = spot.get(ilk_kelime) if ilk_kelime else None

        # GÜN hesabı
        gun = None
        if ikinci_kel and bugun_tarihi:
            vade_sonu = sozlesme.get(ikinci_kel)
            if vade_sonu:
                delta = (vade_sonu - bugun_tarihi).days
                gun = delta if delta > 0 else None

        # HESAPLAMA
        hesaplama = None
        if alis and spot_satis and spot_satis != 0 and gun and gun != 0:
            # ((SATIŞ_spot / ALIŞ_vadeli) - 1) / GÜN * 365
            hesaplama = ((spot_satis / alis) - 1) / gun * 365

        # İŞLEM ÖNERİSİ
        islem_onerisi = None
        if hesaplama is not None and referans_faiz is not None:
            # Karşılaştırma: HESAPLAMA (oran, 0-1 arası) vs FAİZ (yüzde, örn 36.78)
            # Tutarlı olması için ikisini de yüzde cinsine çekelim
            hesaplama_pct = hesaplama * 100
            if hesaplama_pct > referans_faiz:
                islem_onerisi = "İŞLEM YAP"
            else:
                islem_onerisi = "İŞLEM YAPMA"
        
        gun_fark_pct = None
        try:
            val = row.get("GÜN FARK %")
            if val is not None:
                gun_fark_pct = float(val)
        except:
            gun_fark_pct = 0.0

        sonuclar.append({
            "kontrat":        kontrat,
            "aciklama":       aciklama,
            "alis":           alis,
            "spot_satis":     spot_satis,
            "gun":            gun,
            "gun_fark":       gun_fark_pct,
            "hesaplama":      round(hesaplama * 100, 4) if hesaplama is not None else None,
            "referans_faiz":  referans_faiz,
            "islem_onerisi":  islem_onerisi,
        })

    islem_yap_sayi   = sum(1 for r in sonuclar if r["islem_onerisi"] == "İŞLEM YAP")
    islem_yapma_sayi = sum(1 for r in sonuclar if r["islem_onerisi"] == "İŞLEM YAPMA")

    return {
        "meta": {
            "bugun":          str(bugun_tarihi) if bugun_tarihi else None,
            "referans_faiz":  referans_faiz,
            "toplam_satir":   len(sonuclar),
            "islem_yap":      islem_yap_sayi,
            "islem_yapma":    islem_yapma_sayi,
        },
        "sonuclar": sonuclar,
    }

# ─────────────────────────────────────────────────────────
def kaydet_excel(sonuc: dict, output_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb_out = openpyxl.Workbook()
    ws     = wb_out.active
    ws.title = "SONUÇLAR"

    headers = ["KONTRAT", "AÇIKLAMA", "ALIŞ", "SPOT SATIŞ", "GÜN FARK %",
               "HESAPLAMA %", "REFERANS FAİZ %", "İŞLEM ÖNERİSİ"]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    yesil   = PatternFill("solid", fgColor="C6EFCE")
    kirmizi = PatternFill("solid", fgColor="FFC7CE")
    gri     = PatternFill("solid", fgColor="F2F2F2")
    beyaz   = PatternFill("solid", fgColor="FFFFFF")
    thin    = Side(style="thin", color="CCCCCC")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(sonuc["sonuclar"], start=2):
        ws.append([
            r["kontrat"], r["aciklama"], r["alis"], r["spot_satis"],
            r["gun_fark"], r["hesaplama"], r["referans_faiz"], r["islem_onerisi"],
        ])
        fill = gri if i % 2 == 0 else beyaz
        for cell in ws[i]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            cell.border = border

        onerisi_cell = ws.cell(row=i, column=8)
        if r["islem_onerisi"] == "İŞLEM YAP":
            onerisi_cell.fill = yesil
            onerisi_cell.font = Font(bold=True, color="276221")
        elif r["islem_onerisi"] == "İŞLEM YAPMA":
            onerisi_cell.fill = kirmizi
            onerisi_cell.font = Font(bold=True, color="9C0006")

        for col in [3, 4]:
            c = ws.cell(row=i, column=col)
            if c.value is not None:
                c.number_format = "0.00"
        for col in [6, 7]:
            c = ws.cell(row=i, column=col)
            if c.value is not None:
                c.number_format = '0.0000'

    for hdr_cell in ws[1]:
        hdr_cell.border = border

    col_widths = [18, 38, 10, 12, 7, 14, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Özet sheet
    ws2 = wb_out.create_sheet("ÖZET")
    meta = sonuc["meta"]
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 15
    ozet = [
        ("Analiz Tarihi (Bugün)",       meta["bugun"]),
        ("Referans Faiz (TLREF) %",     meta["referans_faiz"]),
        ("Toplam Vadeli Satır",          meta["toplam_satir"]),
        ("İŞLEM YAP",                   meta["islem_yap"]),
        ("İŞLEM YAPMA",                 meta["islem_yapma"]),
    ]
    for row in ozet:
        ws2.append(row)
    for r in range(1, 6):
        ws2.cell(row=r, column=1).font = Font(bold=True)
    ws2.cell(row=4, column=2).fill = yesil
    ws2.cell(row=4, column=2).font = Font(bold=True, color="276221")
    ws2.cell(row=5, column=2).fill = kirmizi
    ws2.cell(row=5, column=2).font = Font(bold=True, color="9C0006")

    wb_out.save(output_path)

# ─────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="FINOVUS ALGO-VAR Hesaplama Motoru")
    parser.add_argument("--input",  required=True, help="FINOVUS.xlsx dosya yolu")
    parser.add_argument("--output", default=None,  help="Sonuç xlsx yolu (opsiyonel)")
    args = parser.parse_args()

    try:
        sonuc = hesapla(args.input)
    except Exception as e:
        print(json.dumps({"hata": str(e)}), file=sys.stderr)
        sys.exit(1)

    if args.output:
        try:
            kaydet_excel(sonuc, args.output)
        except Exception as e:
            print(f"Excel kaydetme hatası: {e}", file=sys.stderr)

    # JSON çıktı → stdout (C# tarafı okur)
    print(json.dumps(sonuc, ensure_ascii=False))

if __name__ == "__main__":
    main()
